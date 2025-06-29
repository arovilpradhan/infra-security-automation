# This Script is used to Format the Raw KPI-1 & KPI-2 vulnerability Full summary report downloaded from Qualys dashboard. 
# This also applies to segregate different excel logic based on Linux & Windows to get a proper Final report as per Linux/Windows/VM Teams requirements.
# This Final report will have 3 Sheets, Raw, KPI-1 & KPI-2. Where Raw Holds all the raw data unfiltered while KPI-1 & 2 have Late/NonLate segments within them,
# Prioritised based on the severity level & impact level.
# === IMPORTS ===
import pandas as pd
from collections import Counter
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import warnings

warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)

# === FILE PATHS ===
kpi1_path = str(input("Enter the qualys kpi-1 detailed report CSV filename(without extension):- ")) + ".csv"
kpi2_path = str(input("Enter the qualys kpi-2 detailed report CSV filename(without extension):- ")) + ".csv"
output_path = str(input("Enter a filename for formatted Excel report.(e.g. Formatted_Qualys_Report):- ")) + ".xlsx"

# === HEADERS ===
summary_header = '"QID","TITLE","CVEIDS","SEVERITY","KB SEVERITY","QDS","DETECTION COUNT"'
details_header = '"QID","Title","Severity","KB Severity","Type Detected","Last Detected","First Detected","Protocol","Port","Status","Asset Id","Asset Name","Asset IPV4","Asset IPV6","Solution","Asset Tags","Disabled","Ignored","QDS","QDS Severity","Detection AGE","Published Date","Patch Released","Category","RTI","Operating System","Last Fixed","Last Reopened","Times Detected","Threat","Vuln Patchable","Asset Critical Score","TruRisk Score","Vulnerability Tags","Results","MITRE ATT&CK TACTIC ID","MITRE ATT&CK TACTIC NAME","MITRE ATT&CK TECHNIQUE ID","MITRE ATT&CK TECHNIQUE NAME"'

from io import StringIO
summary_columns = pd.read_csv(StringIO(summary_header)).columns.tolist()
details_columns = pd.read_csv(StringIO(details_header)).columns.tolist()

# === PARSER ===
def parse_qualys_file(file_path):
    with open(file_path, 'r', encoding='latin1') as f:
        lines = f.readlines()

    summary_rows, details_rows, current_section = [], [], []
    section_type = None

    for line in lines:
        if summary_header in line:
            if current_section and section_type:
                if section_type == 'summary':
                    summary_rows.extend(current_section[1:])
                else:
                    details_rows.extend(current_section[1:])
            current_section = [line]
            section_type = 'summary'
        elif details_header in line:
            if current_section and section_type:
                if section_type == 'summary':
                    summary_rows.extend(current_section[1:])
                else:
                    details_rows.extend(current_section[1:])
            current_section = [line]
            section_type = 'details'
        else:
            current_section.append(line)

    if current_section and section_type:
        if section_type == 'summary':
            summary_rows.extend(current_section[1:])
        else:
            details_rows.extend(current_section[1:])

    summary_df = pd.read_csv(StringIO("".join(summary_rows)), header=None, quotechar='"', encoding='latin1')
    summary_df.columns = summary_columns
    summary_df.dropna(how='all', inplace=True)

    details_df = pd.read_csv(StringIO("".join(details_rows)), header=None, quotechar='"', encoding='latin1')
    details_df.columns = details_columns
    details_df.dropna(how='all', inplace=True)

    return summary_df, details_df

# === PROCESS ===
summary1, details1 = parse_qualys_file(kpi1_path)
summary2, details2 = parse_qualys_file(kpi2_path)
summary_combined = pd.concat([summary1, summary2], ignore_index=True)
raw_combined = pd.concat([details1, details2], ignore_index=True)

# === CLEANUP AND ADD COLUMNS ===
raw_combined.columns = raw_combined.columns.str.strip()
summary_combined.columns = summary_combined.columns.str.strip()
summary_combined.drop(columns=[col for col in ["KB SEVERITY", "QDS", "CVEIDS"] if col in summary_combined.columns], inplace=True)

new_columns = ["Age", "Hostname", "Assignee", "Reasearch", "Action Taken", "Remediation", "Comments"]
insert_at = summary_combined.columns.get_loc("DETECTION COUNT") + 1 if "DETECTION COUNT" in summary_combined.columns else len(summary_combined.columns)
for i, col in enumerate(new_columns):
    summary_combined.insert(insert_at + i, col, "")

# === CALCULATE AGE ===
raw_combined['Title'] = raw_combined['Title'].astype(str).str.strip()
summary_combined['TITLE'] = summary_combined['TITLE'].astype(str).str.strip()
age_results = []
for asset in summary_combined['TITLE']:
    matching_rows = raw_combined[raw_combined['Title'] == asset]
    detection_ages = matching_rows['Detection AGE'].dropna().tolist()
    age_counter = Counter(detection_ages)
    if age_counter:
        most_common_sorted = sorted(age_counter.items(), key=lambda x: (-x[1], -x[0]))
        best_age = most_common_sorted[0][0]
    else:
        best_age = ""
    age_results.append(best_age)
summary_combined['Age'] = age_results

# === SORT BY SEVERITY DESCENDING ===
summary_combined = summary_combined.sort_values(by="SEVERITY", ascending=False)

# === SPLIT INTO KPI-1 AND KPI-2 ===
kpi2_df = summary_combined[summary_combined['SEVERITY'].isin([1, 2, 3])].copy()
kpi1_df = summary_combined[~summary_combined['SEVERITY'].isin([1, 2, 3])].copy()

# === REMOVE SEVERITY COLUMN ===
for df in [kpi1_df, kpi2_df]:
    if 'SEVERITY' in df.columns:
        df.drop(columns=['SEVERITY'], inplace=True)

# === FUNCTION TO SPLIT LATE/NON-LATE ===
def split_late_nonlate(df):
    df['Age'] = pd.to_numeric(df['Age'], errors='coerce')
    late_df = df[df['Age'] > 31]
    nonlate_df = df[df['Age'] <= 31]
    return late_df, nonlate_df

# === SPLIT KPI-1 & KPI-2 ===
kpi1_late, kpi1_nonlate = split_late_nonlate(kpi1_df)
kpi2_late, kpi2_nonlate = split_late_nonlate(kpi2_df)

# === APPLY OS-WISE HOSTNAME MAPPING ===
def os_based_hostname_mapping(df):
    df['Hostname'] = ""

    for i, row in df.iterrows():
        title = row['TITLE']
        raw_matches = raw_combined[raw_combined['Title'].str.strip() == title]

        if not raw_matches.empty:
            os_name = raw_matches['Operating System'].astype(str).str.lower().iloc[0]

            if "red hat" in os_name:
                redhat_matches = raw_matches[raw_matches['Asset Name'].str.contains(r'd\.|d$', na=False, case=False)]
                if not redhat_matches.empty:
                    df.at[i, 'Hostname'] = redhat_matches['Asset Name'].iloc[0]
                else:
                    df.at[i, 'Hostname'] = raw_matches['Asset Name'].iloc[0]

            elif "microsoft windows" in os_name:
                df.drop(columns=['Hostname'], inplace=True)
                break  # Drop for the whole df if it's Windows-based

    return df

kpi1_late = os_based_hostname_mapping(kpi1_late)
kpi1_nonlate = os_based_hostname_mapping(kpi1_nonlate)
kpi2_late = os_based_hostname_mapping(kpi2_late)
kpi2_nonlate = os_based_hostname_mapping(kpi2_nonlate)

# === REMOVE AGE COLUMN BEFORE SAVING ===
for df in [kpi1_late, kpi1_nonlate, kpi2_late, kpi2_nonlate]:
    if 'Age' in df.columns:
        df.drop(columns=['Age'], inplace=True)

# === FUNCTION TO ADD SPACED SECTION TO EXCEL ===
def write_late_nonlate(writer, sheetname, late_df, nonlate_df):
    late_df = late_df.sort_values(by="DETECTION COUNT", ascending=False)
    nonlate_df = nonlate_df.sort_values(by="DETECTION COUNT", ascending=False)

    late_start_row = 0
    late_df.to_excel(writer, sheet_name=sheetname, index=False, startrow=late_start_row + 1)

    ws = writer.sheets[sheetname]
    ws["A1"] = "LATE"
    ws["A1"].font = Font(bold=True)

    nonlate_start_row = late_start_row + len(late_df) + 4 if not nonlate_df.empty else None
    if nonlate_start_row is not None:
        nonlate_df.to_excel(writer, sheet_name=sheetname, index=False, startrow=nonlate_start_row)
        ws[f"A{nonlate_start_row}"] = "NON-LATE"
        ws[f"A{nonlate_start_row}"].font = Font(bold=True)

    for df, start_row in [(late_df, late_start_row + 2), (nonlate_df, nonlate_start_row + 2 if nonlate_start_row is not None else None)]:
        if df.empty:
            continue
        end_row = start_row + len(df)
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# === FINAL FORMATTING ===
def final_formatting(writer):
    for sheetname in ['KPI-1', 'KPI-2', 'Raw']:
        ws = writer.sheets[sheetname]
        
        if sheetname == 'Raw':
            for cell in ws[1]:
                cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        if sheetname in ['KPI-1', 'KPI-2']:
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12

            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                if len(row) >= 4:
                    row[3].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for header_cell in ws[2]:
                header_cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

            if len(ws['A']) > len(kpi1_late) + 4:
                second_section_start_row = len(kpi1_late) + 5
                for header_cell in ws[second_section_start_row]:
                    header_cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

# === SAVE TO FILE ===
if os.path.exists(output_path):
    os.remove(output_path)

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    raw_combined.to_excel(writer, sheet_name='Raw', index=False)
    write_late_nonlate(writer, 'KPI-1', kpi1_late, kpi1_nonlate)
    write_late_nonlate(writer, 'KPI-2', kpi2_late, kpi2_nonlate)
    final_formatting(writer)

print(f"✅ Final report saved: {output_path}")
